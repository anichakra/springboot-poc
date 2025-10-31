import logging
import os
import time

import openpyxl
from langchain_core.messages import SystemMessage, HumanMessage
from langchain_ollama.chat_models import ChatOllama
from langchain_openai import ChatOpenAI
# from langchain_ollama import OllamaLLM
from openpyxl.utils import get_column_letter
from pydantic import SecretStr

MAX_MEMORY_ENTRIES = 5

SYSTEM_PROMPT = "You are an AI assistant, helping in providing insight to camera outputs."

UNIFIED = "unified"

#MODEL = "llava:34b"
MODEL="gemma3:27b"

logger = logging.getLogger("application")

class AnalyticsCallback:
    """
    Class providing analytics callback functionality for processing, logging,
    and saving metadata and results via a language model process.

    This class enables the processing of metadata and frames from grouped cameras.
    It logs results from querying the language model and saves them in
    both a log file and an Excel workbook format. It supports multiple language
    model backends, and maintains thread-specific context for iterative queries.
    Includes capabilities for multithreading and contextual memory to
    process grouped camera data.

    Attributes:
        current_time (float): Stores the current timestamp.
        frame_count (int): The count of frames processed.
        llm: The selected language model object (e.g., ChatOpenAI or ChatOllama).
        _first_callback_time (float): The timestamp of the first callback execution.
        log_wait_time (int): Time interval for logging in seconds.
        current_context (list): Contextual memory for maintaining conversation history.
        prompt (str): The prompt to send to the language model.
        output_path (str): Directory path for saving logs and workbooks.
        log_file_path (str): Full file path for the log file.
        excel_file_path (str): Full file path for the Excel workbook.
    """
    def __init__(self, output_path: str, log_wait_time: int, prompt: str, api_key: str):
        """
        Initializes an instance of a class with configurations for interacting with a language model.
        The class is used to manage specific settings, such as logging, the API key for the model,
        and the prompt that will be used for generating responses. It also sets up paths for output,
        like log and Excel files, and initializes the time tracking variables.

        Args:
            output_path (str): The directory path where output files will be saved.
            log_wait_time (int): Duration in seconds for triggering log details.
            prompt (str): A string containing the text prompt for generating language model responses.
            api_key (str): API key for authenticating with the language model API. If not provided, a
                default model (ChatOllama) is instantiated.

        Raises:
            None
        """
        self.current_time = time.time()
        self.frame_count = 0
        self.llm = None
        self._first_callback_time = time.time()
        self.log_wait_time = log_wait_time
        self.current_context = []

        self.llm = ChatOpenAI(model_name="gpt-4o",
                              temperature=0,
                              openai_api_key=SecretStr(api_key),
                              max_retries=3,
                              max_tokens=500) \
            if api_key else ChatOllama(model=MODEL, temperature=0)

        self.prompt = prompt
        self.output_path = output_path
        self.log_file_path = os.path.join(output_path, "%s.log" % UNIFIED)
        self.excel_file_path = os.path.join(output_path, "%s.xlsx" % UNIFIED)

    def callback(self, message):
        """
        Processes incoming messages by logging and invoking specified handler functions.

        Handles logging of metadata associated with grouped frame data and periodically
        invokes a long-lived model (LLM) process to handle the frames. The handler is
        triggered based on a time interval condition to manage execution rate.

        Args:
            message (dict): Contains grouped frame data, metadata, and associated time.
                - 'frame' (Any): The frame data that needs to be processed.
                - 'grouped_frame_metadata' (str): Metadata associated with the frame in
                  grouped format.
                - 'grouped_time' (float): The time associated with the grouped frame.

        Raises:
            KeyError: If required keys ('frame', 'grouped_frame_metadata', or 'grouped_time')
            are missing in the provided message.
        """
        logger.debug("Message: %s", message["grouped_frame_metadata"])

        if time.time() - self.current_time > self.log_wait_time:
            self.current_time = time.time()
            # Submit the task to the executor for multithreading

            self.invoke_llm(message['frame'], message["grouped_frame_metadata"], message['grouped_time'])

    def invoke_llm(self, combined_frame, metadata, timestamp):

        """
        Invokes the language model (LLM) with the given combined image frame and metadata,
        processes the response, and updates the context and logs.

        This method processes a combined image frame along with metadata of individual images,
        prepares structured prompts for an LLM, invokes the model to generate a response, and
        updates context, logs, and a workbook with results. It also handles any error that occurs
        during execution by logging it.

        Parameters:
            combined_frame: str
                A base64 encoded string representing the combined image data to be provided to
                the LLM as input.
            metadata: List[Dict[str, Any]]
                A list of dictionaries containing metadata for each frame. Each dictionary includes
                details such as 'camera_id', 'frame_number', and 'frame_timestamp'.
            timestamp: Optional[float]
                A timestamp representing the current execution time or context.

        Raises:
            Exception: Occurs if any unexpected error arises during the LLM invocation or
            subsequent operations.
        """
        try:
            grouped_cameras = sorted(set([frame['camera_id'] for frame in metadata]))
            num_cameras = len(grouped_cameras)

            camera_context = [f"Camera {c}" for c in grouped_cameras]

            prompts = [
                SystemMessage(
                    SYSTEM_PROMPT
                ),
                HumanMessage(content=[
                    # {'type': 'text', 'text': "\n".join(self.current_context)},  # Include memory/context
                    {'type': 'text',
                     'text': f"There are {num_cameras} cameras ({', '.join(camera_context)}). {self.prompt}"},
                    {'type': 'image_url', 'image_url': {"url": f"data:image/jpeg;base64,{combined_frame}"}}
                ])
            ]

            response = self.llm.invoke(prompts)

            # Update thread-specific context with response content
            response_text = response.content.strip()

            self.current_context = f"Your response from previous image: {response_text}"

            info = ", ".join(
                f"{frame['camera_id']}-{frame['frame_number']}-{frame['frame_timestamp']}"
                for frame in sorted(metadata, key=lambda x: x['camera_id'])
            )
            self.save_log(info, response_text, timestamp)

            self.save_workbook(info, response_text, timestamp)

            logger.debug(f"LLM Response: {response_text}")

        except Exception as e:
            logger.error(f"An error occurred in process_snapshot: {e}", exc_info=True)

    def save_log(self, info, response_text, timestamp):
        """
        Saves log details to a specified log file. This method appends log information including a
        timestamp, frame details, and an observation response to a log file. It creates the file
        if it doesn't exist already.

        Parameters:
            info (str): The details of the camera capture, including frame number and frame
                timestamp, to be recorded in the log file.
            response_text (str): The observation or response information to be logged.
            timestamp (str): The time of the event to be recorded.

        Raises:
            This method does not explicitly raise any exceptions but may encounter file I/O
            related exceptions such as FileNotFoundError or IOError during execution if the
            log file path is invalid or inaccessible.
        """
        with open(self.log_file_path, "a") as log_file:
            log_file.write(f"Time: {timestamp}\n")
            log_file.write(f"Camera-Frame Number-Frame Timestamp: {info}\n")
            log_file.write(f"Observation: {response_text}\n")
            log_file.write("-" * 50 + "\n")

    def save_workbook(self, info, response_text, timestamp):
        """
        Saves log information into an Excel workbook. If the workbook does not exist,
        it is created with appropriate headers.

        Parameters:
            info (str): The grouped frames or data to log.
            response_text (str): Additional observation or response information to log.
            timestamp (str): The time when the log entry is created, typically in a
                human-readable format.

        Raises:
            None
        """
        # Check if the file exists. If not, create it and add the headers.
        if not os.path.exists(self.excel_file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Logs"
            headers = ["Time", "Grouped Frames", "Observation"]
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
                column_letter = get_column_letter(col_num)
                sheet.column_dimensions[column_letter].width = 20  # Set column width
            workbook.save(self.excel_file_path)
        # Append the new log entry.
        workbook = openpyxl.load_workbook(self.excel_file_path)
        sheet = workbook.active
        sheet.append([timestamp, info, response_text])
        workbook.save(self.excel_file_path)
